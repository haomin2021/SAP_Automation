from openpyxl import load_workbook
import pandas as pd

wb = load_workbook(r"DataLoader\Polierlinie_4_2025.xlsx")
sheet = wb.active

section_tasks = []
current_section = None
pending_task = None

for row in sheet.iter_rows(values_only=True):
    first_cell = row[0]
    second_cell = row[1] if len(row) > 1 else None

    # 标题识别：标题在B列，A列为空
    if first_cell is None and isinstance(second_cell, str) and ":" not in second_cell and len(second_cell.strip()) > 3:
        if pending_task and current_section:
            section_tasks.append((current_section, pending_task))
            pending_task = None
        current_section = second_cell.strip()
        continue

    # 编号行：新任务
    if current_section and isinstance(first_cell, (int, float)):
        if pending_task:
            section_tasks.append((current_section, pending_task))
        desc = str(second_cell).strip() if second_cell else ""
        pending_task = f"{int(first_cell)}. {desc}"
        continue

    # 补充描述行
    if current_section and pending_task and first_cell is None and second_cell:
        pending_task += " " + str(second_cell).strip()

# 最后一条
if pending_task and current_section:
    section_tasks.append((current_section, pending_task))

# 将任务解析为 DataFrame 结构
df = pd.DataFrame(section_tasks, columns=["Section", "FullTask"])

# 拆分编号与描述
df["Task No"] = df["FullTask"].str.extract(r"^(\d+)\.")
df["Description"] = df["FullTask"].str.replace(r"^\d+\.\s*", "", regex=True)
df = df[["Section", "Task No", "Description"]]

# 显示或保存
print(df.head())

# 你也可以保存为 Excel 或 CSV
# df.to_excel("output_tasks.xlsx", index=False)
# df.to_csv("output_tasks.csv", index=False)
