import pandas as pd
from openpyxl import load_workbook

def load_excel(file_path, mode='raw'):
    """
    Load an Excel file and return its content as a DataFrame. Two modes are supported:
    = Mode = 'raw': Directly load the file as a DataFrame. (Default)
    = Mode = 'structured': Load the file and parse it into a structured DataFrame.
        - Title -> Section -> Task Description
    """
    # --------------Mode 'raw'--------------
    if mode == 'raw':
        try:
            return pd.read_excel(file_path, header=None)
        except Exception as e:
            raise RuntimeError("Failed to load Excel file:\n" + str(e))
        
    # --------------Mode 'structured'--------------
    elif mode == 'structured':    
        try:
            wb = load_workbook(file_path)
            sheet = wb.active

            section_tasks = []
            current_section = None
            pending_task = None

            for row in sheet.iter_rows(values_only=True):
                first_cell = row[0]
                second_cell = row[1] if len(row) > 1 else None

                # Title detection: Title in B column, A column is empty
                if first_cell is None and isinstance(second_cell, str) and ":" not in second_cell and len(second_cell.strip()) > 3:
                    if pending_task and current_section:
                        section_tasks.append((current_section, pending_task))
                        pending_task = None
                    current_section = second_cell.strip()
                    continue

                # Identification of numbered rows: new task
                if current_section and isinstance(first_cell, (int, float)):
                    if pending_task:
                        section_tasks.append((current_section, pending_task))
                    desc = str(second_cell).strip() if second_cell else ""
                    pending_task = f"{int(first_cell)}. {desc}"
                    continue

                # Additional description rows
                if current_section and pending_task and first_cell is None and second_cell:
                    pending_task += " " + str(second_cell).strip()

            # Last task handling
            if pending_task and current_section:
                section_tasks.append((current_section, pending_task))

            # Parse tasks into DataFrame structure
            df = pd.DataFrame(section_tasks, columns=["Section", "FullTask"])

            # Parse task number and description
            df["Task No"] = df["FullTask"].str.extract(r"^(\d+)\.")
            df["Description"] = df["FullTask"].str.replace(r"^\d+\.\s*", "", regex=True)
            df = df[["Section", "Task No", "Description"]]

            # Flatten to single column if needed
            lines = []
            current_section = None

            for _, row in df.iterrows():
                section = row["Section"]
                description = str(row["Description"]).strip()

                if section != current_section:
                    lines.append(f"------{section}------")
                    current_section = section

                lines.append(f"{description}")

            return pd.DataFrame(lines, columns=["Task"])
        
        except Exception as e:
            raise RuntimeError("Failed to load structured Excel file:\n" + str(e))
        

        
#################### Example Usage ####################
if __name__ == "__main__":
    # Example usage
    file_path = r"DataLoader\Polierlinie_4_2025.xlsx"
    df = load_excel(file_path, mode='structured')
    print(df.head())